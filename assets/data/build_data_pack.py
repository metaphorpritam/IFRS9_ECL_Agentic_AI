# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "pandas>=2.0",
#   "numpy>=1.24",
#   "scipy>=1.10",
#   "openpyxl>=3.1",
# ]
# ///
"""
build_data_pack.py — builds notes/assets/data/*.xlsx, the Excel data pack for
the IFRS9 study notes. Read notes/plan/conventions.md "Data pack" first.

Run from the repo root:
    uv run --no-project notes/assets/data/build_data_pack.py

(uv resolves the PEP-723 deps above into an isolated env — this script does
NOT touch the project's own --no-sync venv, so it is safe to run even though
openpyxl is not a project dependency.)

Produces, all under notes/assets/data/:
  fixtures_all.xlsx                 — all golden values from tests/fixtures/compute_*.py
  dcr_coefficients.xlsx             — DCR hazard-ratio tables (outputs/hazard/hazard_ratios.md)
  sflld_coefficients.xlsx           — Freddie SFLLD hazard/cure/severity coefficient CSVs
  scenario_weights_calibration.xlsx — DFAST scenario weights, macro anchors, satellite
                                       regression, and the resulting scenario ECL summary

Every sheet: a descriptive title in row 1 (merged-look via a bold header row
written by pandas, not a literal Excel merge — kept simple/robust), a frozen
header row (freeze_panes), and autosized-ish column widths.
"""
from __future__ import annotations

import ast
import importlib.util
import re
import sys
from decimal import Decimal
from pathlib import Path

import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[3]
FIXTURES_DIR = REPO_ROOT / "tests" / "fixtures"
OUT_DIR = Path(__file__).resolve().parent

FIXTURE_MODULES = [
    "compute_ecl", "compute_pd", "compute_vasicek", "compute_scenarios",
    "compute_grossup", "compute_ncl", "compute_rollrate", "compute_validation",
]


# ---------------------------------------------------------------------------
# Excel writing helpers — title row + frozen header + column widths.
# ---------------------------------------------------------------------------
def write_titled_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, title: str) -> None:
    """Write df starting at row 3 (1-indexed) with a title in row 1, freeze
    the header row (row 3, i.e. below the title+blank), and set column widths."""
    sheet_name = sheet_name[:31]  # Excel sheet-name length limit
    df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
    ws = writer.sheets[sheet_name]
    from openpyxl.styles import Font

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    # Freeze panes just below the header row (row 3 is the header -> freeze at A4)
    ws.freeze_panes = "A4"
    for i, col in enumerate(df.columns, start=1):
        try:
            maxlen = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).values[:200]])
        except Exception:
            maxlen = len(str(col))
        ws.column_dimensions[ws.cell(row=3, column=i).column_letter].width = min(60, max(10, maxlen + 2))


# ---------------------------------------------------------------------------
# fixtures_all.xlsx — every golden value from tests/fixtures/compute_*.py
# ---------------------------------------------------------------------------
def _module_source_anchor(source: str) -> str:
    """Pull the docstring's 'Source:' line (or the first non-blank docstring
    line) as the citation anchor for every row from that fixture module."""
    m = re.search(r'"""(.*?)"""', source, re.DOTALL)
    if not m:
        return ""
    doc = m.group(1).strip()
    src_match = re.search(r"Source:\s*(.+?)(?:\n\n|\Z)", doc, re.DOTALL)
    if src_match:
        return " ".join(src_match.group(1).split())
    first_line = doc.splitlines()[0].strip()
    return first_line


def _module_top_level_inputs(source: str) -> str:
    """Extract top-level NAME = <literal> assignments (the stated inputs) via
    ast.literal_eval, stopping once we hit the RESULTS/derivation section.
    Best-effort: skips anything not a plain literal (functions, DataFrames)."""
    tree = ast.parse(source)
    parts = []
    for node in tree.body:
        if isinstance(node, ast.Assign) and len(node.targets) == 1 and isinstance(node.targets[0], ast.Name):
            name = node.targets[0].id
            if name in ("RESULTS", "TARGETS", "_DISPLAY_DECIMALS"):
                break
            try:
                value = ast.literal_eval(node.value)
            except Exception:
                continue
            if isinstance(value, float):
                value = round(value, 6)
            parts.append(f"{name}={value!r}")
    return ", ".join(parts)


def _displayed_decimals(target: float) -> int:
    if float(target).is_integer():
        return 0
    exp = Decimal(str(target)).as_tuple().exponent
    return max(0, -exp) if isinstance(exp, int) else 0


def _import_fixture(mod_name: str):
    spec = importlib.util.spec_from_file_location(
        f"tests.fixtures.{mod_name}", FIXTURES_DIR / f"{mod_name}.py"
    )
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)  # pure numeric derivation, no I/O side effects
    return mod


def build_fixtures_all() -> pd.DataFrame:
    rows = []
    for mod_name in FIXTURE_MODULES:
        path = FIXTURES_DIR / f"{mod_name}.py"
        source = path.read_text(encoding="utf-8")
        anchor = _module_source_anchor(source)
        inputs = _module_top_level_inputs(source)
        mod = _import_fixture(mod_name)
        assert set(mod.RESULTS) == set(mod.TARGETS), f"{mod_name}: RESULTS/TARGETS keys diverge"
        for key, target in mod.TARGETS.items():
            computed = mod.RESULTS[key]
            places = _displayed_decimals(target)
            match = round(float(computed), places) == round(float(target), places)
            rows.append({
                "id": f"{mod_name}::{key}",
                "fixture_module": mod_name,
                "description": key.replace("_", " "),
                "inputs": inputs,
                "computed_value": computed,
                "notes_printed_value": target,
                "displayed_decimals": places,
                "matches_notes": match,
                "source_anchor": anchor,
            })
    df = pd.DataFrame(rows)
    assert df["matches_notes"].all(), (
        "fixtures_all: at least one computed value does not match the notes' "
        "printed value — the golden-value gate (tests/test_fixtures.py) should "
        "already prevent this; investigate before shipping the data pack."
    )
    return df


# ---------------------------------------------------------------------------
# dcr_coefficients.xlsx — outputs/hazard/hazard_ratios.md, two tables
# ---------------------------------------------------------------------------
def _parse_markdown_table(md: str, header_hint: str) -> pd.DataFrame:
    """Parse the first Markdown pipe-table appearing after header_hint."""
    idx = md.index(header_hint)
    tail = md[idx:]
    lines = tail.splitlines()
    table_lines = []
    started = False
    for line in lines:
        if line.strip().startswith("|"):
            started = True
            table_lines.append(line)
        elif started:
            break
    if not table_lines:
        raise ValueError(f"no markdown table found after {header_hint!r}")
    header = [c.strip() for c in table_lines[0].strip().strip("|").split("|")]
    data_lines = [l for l in table_lines[2:]]  # skip header + separator row
    records = []
    for line in data_lines:
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) != len(header):
            continue
        records.append(cells)
    return pd.DataFrame(records, columns=header)


def build_dcr_coefficients() -> dict[str, pd.DataFrame]:
    md_path = REPO_ROOT / "outputs" / "hazard" / "hazard_ratios.md"
    md = md_path.read_text(encoding="utf-8")
    default_hz = _parse_markdown_table(md, "## Default hazard")
    prepay_hz = _parse_markdown_table(md, "## Prepayment hazard")
    return {"default_hazard": default_hz, "prepayment_hazard": prepay_hz}


# ---------------------------------------------------------------------------
# sflld_coefficients.xlsx — Freddie SFLLD hazard/cure/severity CSVs
# ---------------------------------------------------------------------------
def build_sflld_coefficients() -> dict[str, pd.DataFrame]:
    base = REPO_ROOT / "outputs" / "freddie"
    return {
        "hazard_coefficients": pd.read_csv(base / "hazard" / "coefficients.csv"),
        "cure_coefficients": pd.read_csv(base / "lgd" / "cure_coefficients.csv"),
        "severity_coefficients": pd.read_csv(base / "lgd" / "severity_coefficients.csv"),
    }


# ---------------------------------------------------------------------------
# scenario_weights_calibration.xlsx
# ---------------------------------------------------------------------------
def build_scenario_weights_calibration() -> dict[str, pd.DataFrame]:
    scen_md = (REPO_ROOT / "outputs" / "scenarios" / "scenarios_report.md").read_text(encoding="utf-8")
    weights = _parse_markdown_table(scen_md, "## Scenarios and weights")
    macro_anchors = _parse_markdown_table(scen_md, "## Rebasing convention")

    sat_md = (REPO_ROOT / "outputs" / "satellite" / "satellite_report.md").read_text(encoding="utf-8")
    sat_coefs = _parse_markdown_table(sat_md, "## The model")
    sat_headline = pd.DataFrame([{
        "adj_R2": 0.466, "durbin_watson": 1.39, "AIC": 117.05, "n_quarters": 57,
        "model_equation": "Z_t = -1.694 + 13.642*hpi_growth_lag1 + 0.730*gdp_growth_lag2 + e_t",
        "source": "outputs/satellite/satellite_report.md",
    }])

    ecl_summary = pd.read_csv(REPO_ROOT / "outputs" / "scenario_ecl" / "scenario_ecl_summary.csv")

    return {
        "dfast_scenario_weights": weights,
        "macro_anchors": macro_anchors,
        "satellite_model_coefs": sat_coefs,
        "satellite_model_fit": sat_headline,
        "scenario_ecl_summary": ecl_summary,
    }


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------
def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    fixtures_df = build_fixtures_all()
    out_path = OUT_DIR / "fixtures_all.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        write_titled_sheet(writer, fixtures_df, "fixtures_all",
                            "All 133 golden values — tests/fixtures/compute_*.py "
                            "(source of truth for every notes derivation)")
    print(f"wrote {out_path}  ({len(fixtures_df)} rows)")

    dcr = build_dcr_coefficients()
    out_path = OUT_DIR / "dcr_coefficients.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        write_titled_sheet(writer, dcr["default_hazard"], "default_hazard",
                            "DCR default hazard — cloglog hazard ratios (outputs/hazard/hazard_ratios.md)")
        write_titled_sheet(writer, dcr["prepayment_hazard"], "prepayment_hazard",
                            "DCR prepayment hazard — cloglog hazard ratios (outputs/hazard/hazard_ratios.md)")
    print(f"wrote {out_path}  ({sum(len(v) for v in dcr.values())} total rows)")

    sflld = build_sflld_coefficients()
    out_path = OUT_DIR / "sflld_coefficients.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        write_titled_sheet(writer, sflld["hazard_coefficients"], "hazard_coefficients",
                            "SFLLD default hazard coefficients (outputs/freddie/hazard/coefficients.csv)")
        write_titled_sheet(writer, sflld["cure_coefficients"], "cure_coefficients",
                            "SFLLD cure-rate coefficients (outputs/freddie/lgd/cure_coefficients.csv)")
        write_titled_sheet(writer, sflld["severity_coefficients"], "severity_coefficients",
                            "SFLLD severity coefficients (outputs/freddie/lgd/severity_coefficients.csv)")
    print(f"wrote {out_path}  ({sum(len(v) for v in sflld.values())} total rows)")

    scen = build_scenario_weights_calibration()
    out_path = OUT_DIR / "scenario_weights_calibration.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        write_titled_sheet(writer, scen["dfast_scenario_weights"], "dfast_scenario_weights",
                            "DFAST base/down/up scenario construction and weights (outputs/scenarios/scenarios_report.md)")
        write_titled_sheet(writer, scen["macro_anchors"], "macro_anchors",
                            "Macro jump-off levels and long-run reversion targets (outputs/scenarios/scenarios_report.md)")
        write_titled_sheet(writer, scen["satellite_model_coefs"], "satellite_model_coefs",
                            "Satellite regression: credit-cycle Z on macro drivers (outputs/satellite/satellite_report.md)")
        write_titled_sheet(writer, scen["satellite_model_fit"], "satellite_model_fit",
                            "Satellite regression fit statistics (outputs/satellite/satellite_report.md)")
        write_titled_sheet(writer, scen["scenario_ecl_summary"], "scenario_ecl_summary",
                            "Resulting per-scenario and probability-weighted ECL (outputs/scenario_ecl/scenario_ecl_summary.csv)")
    print(f"wrote {out_path}  ({sum(len(v) for v in scen.values())} total rows)")


if __name__ == "__main__":
    main()
